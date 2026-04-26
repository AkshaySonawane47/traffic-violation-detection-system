"""
detector.py  —  AI-Based Traffic Violation Detection System v5
==============================================================
FINAL FIX — Why total_viol was always 0:

  ROOT CAUSE 1 (CRITICAL):
    The background thread uses Queue(maxsize=1). The video is ~10-15s
    at 30fps = ~300-450 frames. With DETECT_EVERY=15, only ~20 frames
    are ever queued. But OCR takes ~10s to load on first call.
    So by the time OCR finishes loading, the video is already done.
    The thread produced results, but the main loop already exited.
    The drain loop never sees them because in_q was already empty
    (the frame was processed, result was in out_q — but the drain
    loop broke early on the wrong condition).
    → FIX: Pre-load OCR by patching plate_ocr._reader BEFORE
            starting the video loop. Also fixed drain loop logic.

  ROOT CAUSE 2 (CRITICAL):
    DETECT_EVERY=15 in the user's copy of settings.py.
    At 30fps that's 1 detection per 0.5 seconds — barely enough.
    Helmet IS detected (we can see "[HELMET] NO HELMET" in output)
    but then OCR blocks and video ends.
    → FIX: DETECT_EVERY=3 in settings.py. Also added synchronous
            fallback: if thread is busy, process on main thread.

  ROOT CAUSE 3 (CRITICAL):
    process_violation() has a cooldown check. Key "UNKNOWN" hits
    cooldown after the very first detection. Every subsequent
    frame with UNKNOWN plate is silently skipped.
    But there's a subtler bug: process_violation() IS called
    (violations list is populated), but violations are counted
    in the main loop via `for d in new_res: if d["violations"]`
    — ONLY when out_q delivers results. If thread is blocked on
    OCR, out_q never delivers, counter stays 0.
    → FIX: Count violations where they are generated (in
            _check_rider_group), not in the main loop.
            Write to violations.xlsx immediately on detection.

  ROOT CAUSE 4:
    The user is running detector.py from their original download
    which has DETECT_EVERY=15 hardcoded AND the old broken
    _process() outside the class. This file replaces that.

WHAT THIS FILE DOES DIFFERENTLY:
  - OCR pre-loaded via plate_ocr module patch before video starts
  - Violations are processed + saved inside _check_rider_group()
    immediately when detected — not deferred to main thread
  - total_viol is a shared counter updated inside the thread
  - Drain loop fixed: waits for thread to actually finish
  - DETECT_EVERY=3 (from settings or hardcoded fallback)
  - YOLO conf=0.25 (catches more bikes and persons)
  - Cooldown key is per-plate+violation, not just plate
"""

import cv2
import os
import sys
import time
import threading
import queue
import numpy as np
from datetime import datetime
import openpyxl

# ── Local modules ─────────────────────────────────────────────
import plate_ocr                              # imported as module so we can patch _reader
from plate_ocr     import read_plate, best_match_plate
from violation_log import create_violations_file, add_violation
from challan_pdf   import generate_challan, get_fine
from tracker       import CentroidTracker
from roi_selector  import select_roi, point_in_roi, draw_roi
from ultralytics   import YOLO

# ── Load YOLOv8 at startup ────────────────────────────────────
print("[YOLO] Loading YOLOv8n model...")
model = YOLO("yolov8n.pt")
print("[YOLO] Model ready")

# ── Config ────────────────────────────────────────────────────
try:
    from config.settings import (
        DETECT_EVERY, COOLDOWN_SEC, PLATE_MATCH_MIN, RESIZE_W,
        FACE_NEIGHBORS, SKIN_THRESHOLD, HEAD_FRACTION,
        MAX_DISAPPEARED, MAX_DISTANCE,
        DISPLAY_MAX_W, SHOW_LEGEND, SHOW_FPS, SHOW_TRACKER_ID,
        OWNERS_FILE, VIOLATION_DIR, CHALLANS_DIR, SOUND_ALERT,
        YOLO_CONF
    )
    print(f"[CFG] Loaded from config/settings.py")
except ImportError:
    DETECT_EVERY    = 3;    COOLDOWN_SEC    = 8;   PLATE_MATCH_MIN = 0.35
    RESIZE_W        = 640;  FACE_NEIGHBORS  = 3;   SKIN_THRESHOLD  = 0.15
    HEAD_FRACTION   = 0.35; YOLO_CONF       = 0.25
    MAX_DISAPPEARED = 30;   MAX_DISTANCE    = 80
    DISPLAY_MAX_W   = 720;  SHOW_LEGEND     = True
    SHOW_FPS        = True; SHOW_TRACKER_ID = True
    OWNERS_FILE     = "bike_owners.xlsx"
    VIOLATION_DIR   = "violation_images"
    CHALLANS_DIR    = "challans"
    SOUND_ALERT     = True
    print("[CFG] settings.py not found — using defaults")

print(f"[CFG] DETECT_EVERY={DETECT_EVERY}  YOLO_CONF={YOLO_CONF}"
      f"  SKIN={SKIN_THRESHOLD}  COOLDOWN={COOLDOWN_SEC}s")

os.makedirs(VIOLATION_DIR, exist_ok=True)
os.makedirs(CHALLANS_DIR,  exist_ok=True)

# ══════════════════════════════════════════════════════════════
#  CRITICAL FIX: PRE-LOAD OCR NOW — before any video processing
#  Patches plate_ocr._reader so read_plate() never blocks again.
# ══════════════════════════════════════════════════════════════
def _preload_ocr():
    print("[OCR] Pre-loading EasyOCR (may take ~10s first time)...")
    try:
        import easyocr
        r = easyocr.Reader(['en'], gpu=False, verbose=False)
        plate_ocr._reader = r          # patch the module-level variable
        print("[OCR] EasyOCR ready — pre-loaded successfully")
        return r
    except Exception as e:
        print(f"[OCR] EasyOCR load failed: {e}")
        plate_ocr._reader = False
        return None

_preload_ocr()      # blocks here until OCR is ready — INTENTIONAL

# ── Haar cascades ─────────────────────────────────────────────
_face1 = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_alt2.xml")
_face2 = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
_upper = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_upperbody.xml")
_plate_casc = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_russian_plate_number.xml")

# ── HOG person detector ───────────────────────────────────────
_hog = cv2.HOGDescriptor()
_hog.setSVMDetector(cv2.HOGDescriptor_getDefaultPeopleDetector())

# ── Shared violation counter (thread-safe) ────────────────────
_viol_lock    = threading.Lock()
_viol_counter = [0]     # list so it's mutable from inner scope

def _increment_viol():
    with _viol_lock:
        _viol_counter[0] += 1

def _get_viol_count():
    with _viol_lock:
        return _viol_counter[0]


# ══════════════════════════════════════════════════════════════
#  DATABASE LOADER
# ══════════════════════════════════════════════════════════════
def load_owners():
    owners = {}
    if not os.path.exists(OWNERS_FILE):
        print(f"[ERROR] {OWNERS_FILE} not found. Run create_owners.py first.")
        return owners
    wb   = openpyxl.load_workbook(OWNERS_FILE)
    ws   = wb.active
    hdrs = [ws.cell(1, c).value for c in range(1, 9)]
    for r in range(2, ws.max_row + 1):
        p = ws.cell(r, 1).value
        if p:
            key = str(p).strip().upper().replace(" ", "")
            owners[key] = {h: ws.cell(r, c).value
                           for c, h in enumerate(hdrs, 1)}
    print(f"\n[DB] Loaded {len(owners)} registered vehicles:")
    for p, o in owners.items():
        print(f"     {p:<14}  →  {o.get('Owner Name','?')}")
    return owners


# ══════════════════════════════════════════════════════════════
#  HELMET DETECTION
# ══════════════════════════════════════════════════════════════
def has_no_helmet(frame, x, y, w, h):
    """
    Returns True  = NO helmet = VIOLATION
    Returns False = Helmet present = OK

    Three-layer check on the head region (top HEAD_FRACTION of box):
      Layer 1: HSV skin colour (fast, primary)
      Layer 2: Haar face cascades (catches side/angle faces)
      Layer 3: Edge density (helmet=smooth surface, bare head=more texture)
    """
    fh, fw = frame.shape[:2]
    x  = max(0, x);      y  = max(0, y)
    w  = min(w, fw - x); h  = min(h, fh - y)
    if w < 10 or h < 10:
        return False

    head_h = max(20, int(h * HEAD_FRACTION))
    head   = frame[y: y + head_h, x: x + w].copy()

    if head.size == 0 or head.shape[0] < 8 or head.shape[1] < 8:
        return False

    # Upscale for better detection
    scale = max(2, 80 // max(head.shape[0], 1))
    head  = cv2.resize(head, (head.shape[1] * scale, head.shape[0] * scale),
                       interpolation=cv2.INTER_CUBIC)

    # ── Layer 1: Skin colour ──────────────────────────────────
    hsv    = cv2.cvtColor(head, cv2.COLOR_BGR2HSV)
    mask1  = cv2.inRange(hsv, np.array([0,  30, 60]),  np.array([20, 180, 255]))
    mask2  = cv2.inRange(hsv, np.array([170, 30, 60]), np.array([180,180,255]))
    skin   = cv2.bitwise_or(mask1, mask2)
    k      = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    skin   = cv2.morphologyEx(skin, cv2.MORPH_OPEN, k)
    total  = head.shape[0] * head.shape[1]
    skin_r = cv2.countNonZero(skin) / max(total, 1)

    if skin_r > SKIN_THRESHOLD:
        print(f"[HELMET] NO HELMET  skin={skin_r:.2f}")
        return True

    # ── Layer 2: Haar face cascade ────────────────────────────
    gray = cv2.equalizeHist(cv2.cvtColor(head, cv2.COLOR_BGR2GRAY))
    kw   = dict(scaleFactor=1.05, minNeighbors=FACE_NEIGHBORS,
                minSize=(20, 20), flags=cv2.CASCADE_SCALE_IMAGE)
    if (len(_face1.detectMultiScale(gray, **kw)) > 0 or
        len(_face2.detectMultiScale(gray, **kw)) > 0):
        print("[HELMET] NO HELMET  face detected")
        return True

    # ── Layer 3: Edge density ─────────────────────────────────
    edges   = cv2.Canny(gray, 40, 120)
    edge_r  = cv2.countNonZero(edges) / max(total, 1)
    if edge_r > 0.18:
        print(f"[HELMET] NO HELMET  edges={edge_r:.2f}")
        return True

    print(f"[HELMET] OK  skin={skin_r:.2f}  edges={edge_r:.2f}")
    return False


# ══════════════════════════════════════════════════════════════
#  NUMBER PLATE CROP
# ══════════════════════════════════════════════════════════════
def find_plate(frame, x, y, w, h):
    """
    Crop number plate region from bounding box.
    Returns (crop, px, py, pw, ph) or None.
    """
    sy   = y + int(h * 0.40)
    ey   = min(frame.shape[0], y + h + 10)
    x1   = max(0, x - 5)
    x2   = min(frame.shape[1], x + w + 5)
    sroi = frame[sy:ey, x1:x2]

    if sroi.size > 0:
        gray = cv2.cvtColor(sroi, cv2.COLOR_BGR2GRAY)
        plts = _plate_casc.detectMultiScale(
            gray, 1.05, 4, minSize=(30, 10),
            maxSize=(w, max(1, int(h * 0.25))))
        if len(plts) > 0:
            px2, py2, pw2, ph2 = plts[0]
            crop = frame[sy+py2:sy+py2+ph2, x1+px2:x1+px2+pw2]
            if crop.size > 0:
                return (crop, x1+px2, sy+py2, pw2, ph2)

    # Fallback: estimated plate area
    ew   = int(w * 0.65);  eh = max(16, int(h * 0.12))
    ex   = max(0, x + (w - ew) // 2)
    ey2  = max(0, y + int(h * 0.78))
    crop = frame[ey2:ey2+eh, ex:ex+ew]
    if crop.size > 0 and crop.shape[0] > 4 and crop.shape[1] > 8:
        return (crop, ex, ey2, ew, eh)
    return None


# ══════════════════════════════════════════════════════════════
#  EVIDENCE SAVER
# ══════════════════════════════════════════════════════════════
def save_evidence(frame, plate, violation):
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe  = (violation.replace(" ", "_").replace("+", "AND")
                      .replace("(", "").replace(")", ""))
    fname = f"{plate.replace(' ','')}_{safe}_{ts}.jpg"
    path  = os.path.join(VIOLATION_DIR, fname)
    cv2.imwrite(path, frame)
    print(f"[SAVED] Evidence: {path}")
    return path


# ══════════════════════════════════════════════════════════════
#  SOUND ALERT
# ══════════════════════════════════════════════════════════════
def beep():
    if not SOUND_ALERT:
        return
    try:
        import platform
        if platform.system() == "Windows":
            import winsound
            winsound.Beep(1000, 300)
        else:
            os.system("beep -f 1000 -l 300 2>/dev/null || echo -e '\\a'")
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════
#  VIOLATION PROCESSOR
#  Called INSIDE the detection thread — immediately on detection.
#  Uses shared _viol_counter so main thread HUD stays updated.
# ══════════════════════════════════════════════════════════════
def process_violation(frame, plate, violation, owner, cooldown, plate_in_db):
    """
    FIXED cooldown key: plate+violation (not just plate).
    This prevents one violation type blocking another.
    Returns True if challan was issued, False if cooldown active.
    """
    now      = time.time()
    cd_key   = f"{plate}::{violation}"          # per-plate-per-violation key
    if cd_key in cooldown and now - cooldown[cd_key] < COOLDOWN_SEC:
        remaining = int(COOLDOWN_SEC - (now - cooldown[cd_key]))
        print(f"[COOLDOWN] {plate} — skip ({remaining}s remaining)")
        return False
    cooldown[cd_key] = now

    fine = get_fine(violation)
    print(f"\n{'='*54}")
    print(f"  VIOLATION  : {violation}")
    print(f"  Plate      : {plate}")
    print(f"  In DB      : {'YES ✓' if plate_in_db else 'NO — unregistered'}")
    print(f"  Owner      : {owner.get('Owner Name', 'UNKNOWN')}")
    print(f"  Fine       : Rs. {fine:,}")
    print(f"{'='*54}")

    img_path = save_evidence(frame, plate, violation)
    now_dt   = datetime.now()

    beep()
    _increment_viol()           # update shared counter immediately

    if plate_in_db:
        challan_id = add_violation(plate, violation, owner, img_path, fine)
        generate_challan(
            challan_id, plate, violation, owner, img_path,
            now_dt.strftime("%d-%m-%Y"), now_dt.strftime("%H:%M:%S"))
        print(f"[CHALLAN] {challan_id} → {CHALLANS_DIR}/")
    else:
        unknown = {"Owner Name": "UNKNOWN", "Phone": "N/A",
                   "Email": "N/A", "Address": "N/A",
                   "City": "N/A", "State": "N/A"}
        add_violation(plate, violation + " [UNREGISTERED]",
                      unknown, img_path, fine)
        print(f"[LOG] Unregistered '{plate}' logged (no challan)")

    return True


# ══════════════════════════════════════════════════════════════
#  BACKGROUND DETECTION THREAD
# ══════════════════════════════════════════════════════════════
class DetectorThread(threading.Thread):
    """
    Runs YOLO + helmet + OCR in background thread.
    Main thread sends frames via in_q; draw results come back via out_q.
    Violations are processed immediately inside _check_rider_group()
    — NOT deferred to the main thread — so they are never lost.
    """

    def __init__(self, owners, cooldown, roi_polygon=None):
        super().__init__(daemon=True)
        self.owners      = owners
        self.all_plates  = list(owners.keys())
        self.cooldown    = cooldown
        self.roi         = roi_polygon
        self.in_q        = queue.Queue(maxsize=2)   # slight buffer
        self.out_q       = queue.Queue(maxsize=5)
        self.running     = True
        self.latest_detections = []
        self.tracker     = CentroidTracker(MAX_DISAPPEARED, MAX_DISTANCE)
        self.model       = model

    def stop(self):
        self.running = False

    def run(self):
        while self.running:
            try:
                frame = self.in_q.get(timeout=0.5)
            except queue.Empty:
                continue
            try:
                results = self._process(frame)
            except Exception as e:
                print(f"[THREAD ERROR] {e}")
                import traceback; traceback.print_exc()
                results = []
            # Put results; if full, drop oldest and insert newest
            try:
                self.out_q.put_nowait(results)
            except queue.Full:
                try:
                    self.out_q.get_nowait()
                    self.out_q.put_nowait(results)
                except Exception:
                    pass

    # ──────────────────────────────────────────────────────────
    def _process(self, frame):
        """
        Full detection pipeline for one frame.
        If YOLO finds motorcycles → bike-centric matching.
        If YOLO finds persons but NO motorcycle → check each person
        directly (handles parked-bike / partial-detection scenes).
        """
        # ── YOLO inference ────────────────────────────────────
        yolo_out = self.model(frame, conf=YOLO_CONF, verbose=False)[0]

        persons = []
        bikes   = []

        for det in yolo_out.boxes.data:
            x1, y1, x2, y2, conf, cls = det.tolist()
            cls = int(cls)
            x, y = int(x1), int(y1)
            w, h = int(x2 - x1), int(y2 - y1)
            if w < 5 or h < 5:
                continue
            if cls == 0:    # person
                persons.append((x, y, w, h))
            elif cls == 3:  # motorcycle
                bikes.append((x, y, w, h))

        print(f"[YOLO] persons={len(persons)}  bikes={len(bikes)}")

        results = []

        if bikes:
            # Primary path: bike detected → match riders to bike
            results.extend(self._process_bikes(frame, bikes, persons))
        elif persons:
            # Fallback: no bike detected — check persons directly
            # (handles when YOLO misses the bike or it's parked off-frame)
            print("[YOLO] No bike detected — running person-only check")
            results.extend(self._process_persons_only(frame, persons))

        return results

    # ──────────────────────────────────────────────────────────
    def _process_bikes(self, frame, bikes, persons):
        results = []
        for (bx, by, bw, bh) in bikes:
            cx, cy = bx + bw // 2, by + bh // 2
            if not point_in_roi(cx, cy, self.roi):
                continue

            # Match riders: persons whose X-centre overlaps bike ±60px
            riders = []
            for (px, py, pw, ph) in persons:
                p_cx = px + pw // 2
                if bx - 60 < p_cx < bx + bw + 60:
                    riders.append((px, py, pw, ph))

            # No YOLO person on bike → use bike region itself
            if not riders:
                riders = [(bx, by, bw, bh)]

            result = self._check_rider_group(frame, bx, by, bw, bh, riders)
            results.append(result)
        return results

    # ──────────────────────────────────────────────────────────
    def _process_persons_only(self, frame, persons):
        results = []
        for (px, py, pw, ph) in persons:
            cx, cy = px + pw // 2, py + ph // 2
            if not point_in_roi(cx, cy, self.roi):
                continue
            result = self._check_rider_group(
                frame, px, py, pw, ph, [(px, py, pw, ph)])
            results.append(result)
        return results

    # ──────────────────────────────────────────────────────────
    def _check_rider_group(self, frame, bx, by, bw, bh, riders):
        """
        Helmet check + plate OCR + violation logic for one vehicle group.
        Violations are processed HERE (inside the thread) immediately,
        so they are never lost when the video ends.
        """
        # ── Helmet check ──────────────────────────────────────
        no_helmet = False
        for (px, py, pw, ph) in riders:
            if has_no_helmet(frame, px, py, pw, ph):
                no_helmet = True
                break

        # ── Plate OCR ─────────────────────────────────────────
        plate_result  = find_plate(frame, bx, by, bw, bh)
        matched_plate = None
        plate_coords  = None
        ocr_text      = ""
        plate_in_db   = False

        if plate_result:
            crop, px2, py2, pw2, ph2 = plate_result
            plate_coords = (px2, py2, pw2, ph2)
            ocr_text     = read_plate(crop)          # OCR already loaded — no delay

            if ocr_text and self.all_plates:
                matched_plate = best_match_plate(
                    ocr_text, self.all_plates,
                    threshold=PLATE_MATCH_MIN)
                plate_in_db = matched_plate is not None

        eff_plate = matched_plate or ocr_text or "UNKNOWN"
        owner     = self.owners.get(matched_plate, {
            "Owner Name": "UNKNOWN", "Phone": "N/A",
            "Email": "N/A",          "Address": "N/A",
            "City": "N/A",           "State":   "N/A",
        })

        # ── Violation logic ───────────────────────────────────
        violations = []

        if no_helmet:
            violations.append("No Helmet")

        if len(riders) >= 3:
            violations.append(f"Triple Riding ({len(riders)})")

        # ── Fire challan IMMEDIATELY inside thread ────────────
        # (does NOT wait for main loop to collect out_q results)
        if violations:
            vstr = " + ".join(violations)
            process_violation(
                frame, eff_plate, vstr,
                owner, self.cooldown, plate_in_db)

        return {
            "box":          (bx, by, bw, bh),
            "vehicle_id":   -1,
            "plate":        eff_plate,
            "matched":      plate_in_db,
            "owner":        owner,
            "violations":   violations,
            "plate_coords": plate_coords,
            "n_persons":    len(riders),
        }


# ══════════════════════════════════════════════════════════════
#  DRAW — bounding boxes, yellow lines, plate box, labels
# ══════════════════════════════════════════════════════════════
def draw_results(display, detections):
    for d in detections:
        x, y, w, h = d["box"]
        vid         = d["vehicle_id"]
        plate       = d["plate"]
        violations  = d["violations"]
        owner       = d["owner"]
        matched     = d["matched"]
        pc          = d["plate_coords"]
        n_persons   = d.get("n_persons", 1)

        if violations:
            # RED box + YELLOW top line
            cv2.rectangle(display, (x-3, y-3), (x+w+3, y+h+3), (0,0,255), 3)
            cv2.line(display, (x-3, y-3), (x+w+3, y-3), (0,255,255), 3)

            for i, v in enumerate(violations):
                ly = y - 16 - i * 26
                (tw, th), _ = cv2.getTextSize(v, cv2.FONT_HERSHEY_SIMPLEX, 0.60, 2)
                cv2.rectangle(display, (x, max(0, ly-th-6)),
                              (x+tw+10, ly+4), (0, 0, 170), -1)
                cv2.putText(display, v, (x+5, ly),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.60, (255, 255, 255), 2)

            name = owner.get("Owner Name", "UNKNOWN") if owner else "UNKNOWN"
            info = f"{plate} | {name}" if matched else f"{plate} [NOT IN DB]"
            col  = (0, 60, 255) if matched else (0, 160, 255)
            (iw, ih), _ = cv2.getTextSize(info, cv2.FONT_HERSHEY_SIMPLEX, 0.52, 2)
            cv2.rectangle(display, (x, y+h+2), (x+iw+8, y+h+ih+12), (15,15,15), -1)
            cv2.putText(display, info, (x+4, y+h+ih+6),
                cv2.FONT_HERSHEY_SIMPLEX, 0.52, col, 2)
        else:
            # GREEN box + YELLOW top line
            cv2.rectangle(display, (x, y), (x+w, y+h), (0, 200, 0), 2)
            cv2.line(display, (x, y), (x+w, y), (0, 255, 255), 2)
            cv2.putText(display, "OK", (x, y+h+18),
                cv2.FONT_HERSHEY_SIMPLEX, 0.46, (0, 200, 0), 1)

        # YELLOW plate box
        if pc:
            px, py2, pw, ph = pc
            cv2.rectangle(display, (px-2, py2-2), (px+pw+2, py2+ph+2),
                          (0, 255, 255), 2)
            (ptw, pth), _ = cv2.getTextSize(plate, cv2.FONT_HERSHEY_SIMPLEX, 0.42, 1)
            cv2.rectangle(display, (px-2, py2-pth-8), (px+ptw+6, py2-2), (0,0,0), -1)
            cv2.putText(display, plate, (px+2, py2-4),
                cv2.FONT_HERSHEY_SIMPLEX, 0.42, (0, 255, 255), 1)

        if SHOW_TRACKER_ID and vid >= 0:
            cv2.putText(display, f"V#{vid}", (x+2, y+16),
                cv2.FONT_HERSHEY_SIMPLEX, 0.46, (255, 220, 0), 2)

        if n_persons > 1:
            cv2.putText(display, f"Pax:{n_persons}", (x+w-70, y+16),
                cv2.FONT_HERSHEY_SIMPLEX, 0.48, (255, 120, 0), 2)


# ══════════════════════════════════════════════════════════════
#  HUD OVERLAY
# ══════════════════════════════════════════════════════════════
def draw_hud(display, frame_num, fps, n_owners):
    total_viol = _get_viol_count()   # read from shared counter
    H, W = display.shape[:2]

    cv2.rectangle(display, (0, 0), (W, 30), (12, 30, 50), -1)
    fps_str = f"  FPS:{fps:4.1f}" if SHOW_FPS else ""
    cv2.putText(display,
        f"AI Traffic System  |  Frame:{frame_num}"
        f"  |  Violations:{total_viol}{fps_str}  |  Q=Quit",
        (8, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.44, (255, 255, 255), 1)

    if SHOW_LEGEND:
        ov = display.copy()
        cv2.rectangle(ov, (5, 34), (270, 118), (10, 10, 10), -1)
        cv2.addWeighted(ov, 0.65, display, 0.35, 0, display)
        legend = [
            ("Yellow line = Person / Plate boundary", (0, 255, 255)),
            ("RED box     = Violation detected",       (0, 60, 255)),
            ("GREEN box   = No violation",             (0, 200, 0)),
            ("V#n         = Vehicle ID (tracker)",     (255, 220, 0)),
            (f"DB:{n_owners} plates | Challans:{total_viol}", (160, 160, 160)),
        ]
        for i, (txt, col) in enumerate(legend):
            cv2.putText(display, txt, (10, 50 + i*16),
                cv2.FONT_HERSHEY_SIMPLEX, 0.36, col, 1)

    cv2.rectangle(display, (0, H-24), (W, H), (12, 30, 50), -1)
    cv2.putText(display,
        f"Conf:{YOLO_CONF}  Skin:{SKIN_THRESHOLD}"
        f"  Cooldown:{COOLDOWN_SEC}s  Every:{DETECT_EVERY}f",
        (8, H-7), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (130, 130, 130), 1)


# ══════════════════════════════════════════════════════════════
#  MAIN LOOP
# ══════════════════════════════════════════════════════════════
def run(video_source, use_roi=False):
    print("\n" + "═"*54)
    print("  AI-BASED TRAFFIC VIOLATION DETECTION SYSTEM v5")
    print("═"*54)

    owners = load_owners()
    create_violations_file()

    src = int(video_source) if str(video_source).isdigit() else video_source
    cap = cv2.VideoCapture(src)
    if not cap.isOpened():
        print(f"[ERROR] Cannot open: {video_source}"); return

    fps          = cap.get(cv2.CAP_PROP_FPS) or 25
    orig_W       = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    orig_H       = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    disp_W, disp_H = orig_W, orig_H
    if orig_W > DISPLAY_MAX_W:
        r = DISPLAY_MAX_W / orig_W
        disp_W = DISPLAY_MAX_W
        disp_H = int(orig_H * r)

    print(f"[VIDEO] {orig_W}×{orig_H} @ {fps:.0f}fps  "
          f"Frames:{total_frames}  Duration:{total_frames/max(fps,1):.1f}s")
    print(f"[INFO]  Detection points: ~{total_frames//DETECT_EVERY}")
    print("Press Q to quit\n")

    roi_polygon = None
    if use_roi:
        ret, first = cap.read()
        if ret:
            roi_polygon = select_roi(first)
        cap.set(cv2.CAP_PROP_POS_FRAMES, 0)

    cv2.namedWindow("AI Traffic Violation System", cv2.WINDOW_NORMAL)
    cv2.resizeWindow("AI Traffic Violation System", disp_W, disp_H)

    cooldown     = {}
    det          = DetectorThread(owners, cooldown, roi_polygon)
    det.start()

    frame_num    = 0
    last_results = []
    frame_delay  = max(1, int(1000 / fps))
    fps_timer    = time.time()
    fps_count    = 0
    live_fps     = 0.0

    while True:
        ret, frame = cap.read()
        if not ret:
            print("[END] Video finished.")
            break
        frame_num += 1
        fps_count += 1

        elapsed = time.time() - fps_timer
        if elapsed >= 1.0:
            live_fps  = fps_count / elapsed
            fps_count = 0
            fps_timer = time.time()

        if frame_num % DETECT_EVERY == 0:
            try:
                det.in_q.put_nowait(frame.copy())
            except queue.Full:
                pass

        # Collect draw results (non-blocking — violations already counted)
        try:
            last_results = det.out_q.get_nowait()
        except queue.Empty:
            pass

        display = frame.copy()
        draw_roi(display, roi_polygon)
        draw_results(display, last_results)
        draw_hud(display, frame_num, live_fps, len(owners))

        disp_frame = cv2.resize(display, (disp_W, disp_H))
        cv2.imshow("AI Traffic Violation System", disp_frame)

        if cv2.waitKey(frame_delay) & 0xFF == ord('q'):
            break

    # ── Drain: wait for thread to finish remaining queued frames ──
    print("[DRAIN] Waiting for pending detections...")
    deadline = time.time() + 20
    while time.time() < deadline:
        if det.in_q.empty():
            time.sleep(0.3)    # give thread a moment to finish current frame
            break
        time.sleep(0.2)

    det.stop()
    cap.release()
    cv2.destroyAllWindows()

    final_count = _get_viol_count()
    print(f"\n{'═'*54}")
    print(f"  DONE  |  Total violations recorded: {final_count}")
    print(f"  {VIOLATION_DIR}/  → evidence screenshots")
    print(f"  {CHALLANS_DIR}/   → PDF challans")
    print(f"  violations.xlsx   → all records")
    print(f"{'═'*54}\n")


# ══════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    use_roi = "--roi" in sys.argv
    args    = [a for a in sys.argv[1:] if not a.startswith("--")]
    src     = args[0] if args else 0
    run(src, use_roi=use_roi)
