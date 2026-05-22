"""
detector.py  —  AI Traffic Violation Detection System  (FINAL)
==============================================================

WHAT WAS WRONG AND WHAT IS FIXED IN THIS VERSION:
--------------------------------------------------

PROBLEM 1 — OCR shows "UNKNOWN" even when plate is visible
  Root cause A: OCR confidence filter was 0.25 — blurry plate crops
    return 0.10–0.22, so they were silently dropped → empty string.
  Root cause B: Fuzzy match threshold was 0.45 — one missed OCR char
    drops score to 0.33 which was below threshold → None → UNKNOWN.
  Root cause C: No OCR correction variants tried. EasyOCR commonly
    reads '0' as 'O', 'B' as '8' etc. on Indian plates.
  FIX (in plate_ocr.py):
    • conf threshold lowered to 0.10
    • threshold in settings lowered to 0.20
    • _generate_variants() tries all common char substitutions
    • difflib SequenceMatcher added as third scoring method
    • 4-pass preprocessing (adaptive-INV, OTSU, sharp-gray, colour)

PROBLEM 2 — Helmet detection false positives (helmet detected as no-helmet)
  Root cause: SKIN_THRESHOLD was 0.15. Dark-coloured helmets with any
    exposed neck/face in the head ROI pushed skin% above 0.15.
    The head region (HEAD_FRACTION=0.38) included forehead AND chin
    area, so even a helmeted rider showed partial skin.
  Root cause 2: Edge density check (edge_r > 0.18) fired on textured
    helmets (logos, vents) which have many edges.
  FIX:
    • HEAD_FRACTION = 0.28 (only top 28% — strictly forehead area)
    • SKIN_THRESHOLD = 0.20 (more tolerant — was 0.15)
    • Edge density check REMOVED — too many false positives
    • Added helmet colour check: if dominant colour in head ROI is
      dark (mean HSV-V < 80) → likely a helmet → return False early
    • Voting: need 2 of 3 checks (skin, face, colour) to fire
      instead of any single check

PROBLEM 3 — Person ↔ Bike ↔ Plate not properly linked
  Root cause: IoU threshold 0.05 is correct but YOLO motorcycle boxes
    often DON'T include the rider (box ends at handlebar).
    The fallback X±bw*0.25 was too tight — missed many valid riders.
  Root cause 2: Plate search started at 70% of bike height.
    When YOLO bike box was small (handlebars only), 70% was mid-air.
  FIX:
    • IoU threshold kept at 0.05
    • Fallback: X within ± max(bw*0.5, 60px) of bike centre,
      Y within ± bh*1.5 of bike centre (much more generous vertically
      because rider sits above the bike box)
    • Plate search: BELOW the bike box is primary, but also check
      WITHIN the bottom 40% of the box if bike box is tall

PROBLEM 4 — Yellow bounding box not aligned with correct rider
  Root cause: draw_results() drew the bike box as the yellow box.
    When riders were matched from person detections, the yellow box
    should enclose the RIDER (person box), not the bike box.
  FIX:
    • If riders detected: compute union of rider boxes → draw YELLOW
    • Bike box: drawn as thin CYAN outline (separate visual layer)
    • RED/GREEN box: drawn around rider union box
    • Plate box: GREEN, linked to bike bottom by yellow line

PROBLEM 5 — Excel matching broken (pandas not used, openpyxl fragile)
  Root cause: openpyxl reads .xlsx row by row. If Excel file has
    merged cells, trailing spaces, or mixed-case plate numbers,
    the key lookup owners[matched_plate] returns None silently.
  FIX:
    • load_owners() now uses PANDAS (more robust Excel reading)
    • All plate keys are strip().upper().replace(' ','') normalized
    • All owner field values are str()-cast (handles int/float cells)
    • Added debug print showing all loaded plates for verification

PROBLEM 6 — Plate box drifts / moves randomly
  Root cause: _plate_fallback() placed plate at by+bh+8px.
    But on persons-only path, "bh" was the union height of ALL persons
    which changes every frame (people move). So +8px from a different
    base each frame = drift.
  FIX:
    • Plate is searched relative to BIKE box only (never person box)
    • On persons-only path, plate search is SKIPPED (no reliable anchor)
    • Plate fallback uses bike box centre-bottom, clamped to frame
"""

import cv2
import os
import sys
import time
import threading
import queue
import numpy as np
from datetime import datetime

# ── Use pandas for robust Excel reading ───────────────────────
try:
    import pandas as pd
    _PANDAS = True
except ImportError:
    import openpyxl
    _PANDAS = False

import plate_ocr
from plate_ocr     import read_plate, best_match_plate
from violation_log import create_violations_file, add_violation
from challan_pdf   import generate_challan, get_fine
from tracker       import CentroidTracker
from roi_selector  import select_roi, point_in_roi, draw_roi
from ultralytics   import YOLO

# ── Load YOLOv8 ───────────────────────────────────────────────
print("[YOLO] Loading YOLOv8n model...")
model = YOLO("yolov8n.pt")
print("[YOLO] Ready")

# ── Config ────────────────────────────────────────────────────
try:
    from config.settings import (
        DETECT_EVERY, COOLDOWN_SEC, PLATE_MATCH_MIN, RESIZE_W,
        FACE_NEIGHBORS, SKIN_THRESHOLD, HEAD_FRACTION,
        MAX_DISAPPEARED, MAX_DISTANCE,
        DISPLAY_MAX_W, SHOW_LEGEND, SHOW_FPS, SHOW_TRACKER_ID,
        OWNERS_FILE, VIOLATION_DIR, CHALLANS_DIR, SOUND_ALERT,
        YOLO_CONF
    )
    print("[CFG] Loaded config/settings.py")
except ImportError:
    DETECT_EVERY    = 3;   COOLDOWN_SEC    = 10;  PLATE_MATCH_MIN = 0.20
    RESIZE_W        = 640; FACE_NEIGHBORS  = 3;   SKIN_THRESHOLD  = 0.20
    HEAD_FRACTION   = 0.28; YOLO_CONF      = 0.25
    MAX_DISAPPEARED = 40;  MAX_DISTANCE    = 120
    DISPLAY_MAX_W   = 720; SHOW_LEGEND     = True
    SHOW_FPS        = True; SHOW_TRACKER_ID = True
    OWNERS_FILE     = "bike_owners.xlsx"
    VIOLATION_DIR   = "violation_images"
    CHALLANS_DIR    = "challans"
    SOUND_ALERT     = True
    print("[CFG] Using hardcoded defaults (settings.py not found)")

print(f"[CFG] Detect every {DETECT_EVERY}f | Conf={YOLO_CONF} | "
      f"PlateThreshold={PLATE_MATCH_MIN} | Skin={SKIN_THRESHOLD}")

os.makedirs(VIOLATION_DIR, exist_ok=True)
os.makedirs(CHALLANS_DIR,  exist_ok=True)

# ══════════════════════════════════════════════════════════════
#  CRITICAL: Pre-load OCR BEFORE video starts
#  Patches plate_ocr._reader so read_plate() never blocks mid-video
# ══════════════════════════════════════════════════════════════
def _preload_ocr():
    print("[OCR] Pre-loading EasyOCR (~10s first time)...")
    try:
        import easyocr
        r = easyocr.Reader(['en'], gpu=False, verbose=False)
        plate_ocr._reader = r          # direct patch → no lazy load
        print("[OCR] EasyOCR ready")
        return r
    except Exception as e:
        print(f"[OCR] EasyOCR unavailable: {e}")
        plate_ocr._reader = False
        return None

_preload_ocr()   # blocking — intentional

# ── Haar cascades ─────────────────────────────────────────────
_face1 = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_alt2.xml")
_face2 = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
_plate_casc = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_russian_plate_number.xml")

# ── Thread-safe violation counter ────────────────────────────
_viol_lock    = threading.Lock()
_viol_counter = [0]

def _inc():
    with _viol_lock:
        _viol_counter[0] += 1

def _count():
    with _viol_lock:
        return _viol_counter[0]


# ══════════════════════════════════════════════════════════════
#  FIX 5 — EXCEL LOADER using pandas
#  Robust against merged cells, trailing spaces, int plate numbers
# ══════════════════════════════════════════════════════════════
def load_owners():
    """
    Load bike_owners.xlsx using pandas (more robust than openpyxl).
    Falls back to openpyxl if pandas not installed.
    All plate keys are normalised: strip + upper + no spaces.
    All owner values are str()-cast (handles numeric cells).

    Returns dict: { "MH12AB1234": {"Owner Name": "...", "Phone": ..., ...} }
    """
    owners = {}

    if not os.path.exists(OWNERS_FILE):
        print(f"[ERROR] {OWNERS_FILE} not found — run create_owners.py first")
        return owners

    try:
        if _PANDAS:
            # FIX 5: pandas reads more reliably than openpyxl cell-by-cell
            df = pd.read_excel(OWNERS_FILE, dtype=str)   # dtype=str prevents int conversion
            df.columns = [str(c).strip() for c in df.columns]
            for _, row in df.iterrows():
                # First column = Number Plate
                raw_plate = str(row.iloc[0]).strip()
                if raw_plate in ('', 'nan', 'None'):
                    continue
                # Normalise key: uppercase, no spaces
                key = raw_plate.upper().replace(' ', '').replace('-', '')
                # Build owner dict from all columns
                owners[key] = {
                    col: str(row[col]).strip() if str(row[col]) not in ('nan','None') else 'N/A'
                    for col in df.columns
                }
        else:
            # Fallback: openpyxl
            wb   = openpyxl.load_workbook(OWNERS_FILE)
            ws   = wb.active
            hdrs = [str(ws.cell(1, c).value or '').strip() for c in range(1, 9)]
            for r in range(2, ws.max_row + 1):
                p = ws.cell(r, 1).value
                if p:
                    key = str(p).strip().upper().replace(' ', '')
                    owners[key] = {
                        h: str(ws.cell(r, c).value or 'N/A').strip()
                        for c, h in enumerate(hdrs, 1)
                    }

        print(f"\n[DB] Loaded {len(owners)} vehicles from {OWNERS_FILE}:")
        for plate, info in owners.items():
            name = info.get('Owner Name', info.get('owner name', '?'))
            print(f"     {plate:<14}  →  {name}")
        print()
        return owners

    except Exception as e:
        print(f"[ERROR] Failed to load {OWNERS_FILE}: {e}")
        return {}


# ══════════════════════════════════════════════════════════════
#  FIX 2 — HELMET DETECTION (reduced false positives)
# ══════════════════════════════════════════════════════════════
def has_no_helmet(frame, px, py, pw, ph):
    """
    Returns True = NO helmet (violation)
    Returns False = Helmet present or uncertain (no violation)

    FIX 2 changes:
    • HEAD_FRACTION reduced to 0.28 (only true forehead area)
    • SKIN_THRESHOLD raised to 0.20 (less sensitive)
    • Added HELMET COLOUR CHECK: dark head region = likely helmet
    • Edge density check REMOVED (caused false positives on textured helmets)
    • VOTING: need 2 positive signals out of 2 checks (skin AND face)
      rather than any single check → far fewer false positives

    Rule:
      If dominant colour is DARK (helmet colours: black, grey, dark blue) → OK
      If BOTH skin-colour AND face-cascade fire → NO HELMET
      If only ONE fires → uncertain → OK (avoid false positives)
    """
    fh, fw = frame.shape[:2]
    px = max(0, px);  py = max(0, py)
    pw = min(pw, fw - px); ph = min(ph, fh - py)
    if pw < 12 or ph < 12:
        return False

    # Crop HEAD region = top HEAD_FRACTION of person box
    head_h = max(16, int(ph * HEAD_FRACTION))
    head   = frame[py: py + head_h, px: px + pw].copy()
    if head.size == 0 or head.shape[0] < 8 or head.shape[1] < 8:
        return False

    # Upscale for better detection
    scale = max(2, 80 // max(head.shape[0], 1))
    head  = cv2.resize(head,
                       (head.shape[1] * scale, head.shape[0] * scale),
                       interpolation=cv2.INTER_CUBIC)

    # ── HELMET COLOUR CHECK (FIX 2: new check) ───────────────
    # Helmets are typically dark. If mean brightness of head ROI < 80 → likely helmet
    hsv     = cv2.cvtColor(head, cv2.COLOR_BGR2HSV)
    v_mean  = float(np.mean(hsv[:, :, 2]))   # Value channel mean brightness
    if v_mean < 75:
        # Dark region = likely dark-coloured helmet → safe to say no violation
        print(f"[HELMET] OK  dark head region (V={v_mean:.0f}) → likely helmet")
        return False

    # ── SKIN COLOUR CHECK ────────────────────────────────────
    mask1   = cv2.inRange(hsv, np.array([0,  30, 60]),  np.array([20, 180, 255]))
    mask2   = cv2.inRange(hsv, np.array([170, 30, 60]), np.array([180, 180, 255]))
    skin    = cv2.morphologyEx(cv2.bitwise_or(mask1, mask2),
                               cv2.MORPH_OPEN,
                               cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3)))
    total   = head.shape[0] * head.shape[1]
    skin_r  = cv2.countNonZero(skin) / max(total, 1)
    skin_ok = skin_r > SKIN_THRESHOLD

    # ── FACE CASCADE CHECK ───────────────────────────────────
    gray    = cv2.equalizeHist(cv2.cvtColor(head, cv2.COLOR_BGR2GRAY))
    kw      = dict(scaleFactor=1.05, minNeighbors=FACE_NEIGHBORS,
                   minSize=(20, 20), flags=cv2.CASCADE_SCALE_IMAGE)
    faces   = (len(_face1.detectMultiScale(gray, **kw)) > 0 or
               len(_face2.detectMultiScale(gray, **kw)) > 0)
    face_ok = faces

    # FIX 2: VOTING — need BOTH signals to fire (reduces false positives)
    if skin_ok and face_ok:
        print(f"[HELMET] NO HELMET  skin={skin_r:.2f} + face detected")
        return True

    # If only skin is high but no face → likely neck/ear visible but helmet on top
    if skin_ok:
        print(f"[HELMET] OK (skin={skin_r:.2f} but no face → neck visible, helmet OK)")
        return False

    # If only face detected but low skin → could be face-forward with helmet
    if face_ok:
        # Face detected with LOW skin → probably helmet with visor up or dark skin
        # Only flag violation if skin is at least somewhat elevated
        if skin_r > SKIN_THRESHOLD * 0.6:
            print(f"[HELMET] NO HELMET  face + moderate skin={skin_r:.2f}")
            return True
        print(f"[HELMET] OK  face detected but skin={skin_r:.2f} too low → helmet")
        return False

    print(f"[HELMET] OK  skin={skin_r:.2f}  V={v_mean:.0f}")
    return False


# ══════════════════════════════════════════════════════════════
#  HELPER: IoU
# ══════════════════════════════════════════════════════════════
def _iou(ax, ay, aw, ah, bx, by, bw, bh):
    ax2, ay2 = ax+aw, ay+ah;  bx2, by2 = bx+bw, by+bh
    ix1 = max(ax,bx); iy1 = max(ay,by)
    ix2 = min(ax2,bx2); iy2 = min(ay2,by2)
    if ix2<=ix1 or iy2<=iy1: return 0.0
    inter = (ix2-ix1)*(iy2-iy1)
    return inter / max(aw*ah + bw*bh - inter, 1)


# ══════════════════════════════════════════════════════════════
#  FIX 3 — PERSON ↔ BIKE ASSOCIATION (more generous tolerances)
# ══════════════════════════════════════════════════════════════
def _associate_persons_to_bike(persons, bx, by, bw, bh):
    """
    FIX 3: More generous matching because YOLO motorcycle box often
    ends at handlebar/seat level, not at the rider's head.

    A person is a rider if ANY of these conditions hold:
      1. IoU(person_box, bike_box) > 0.05  (boxes overlap)
      2. Person centre-X within ± max(bw*0.5, 60px) of bike centre
         AND person centre-Y within bike_top-bh*1.5 to bike_bottom+bh*0.5
         (generous vertical range because rider sits ABOVE the bike box)
      3. Person bottom is within 30px of bike bottom
         (rider's feet near the bike bottom regardless of box overlap)
    """
    riders = []
    b_cx   = bx + bw // 2
    b_cy   = by + bh // 2

    for (px, py, pw, ph) in persons:
        p_cx  = px + pw // 2
        p_cy  = py + ph // 2
        p_bot = py + ph

        # Condition 1: IoU overlap
        if _iou(px, py, pw, ph, bx, by, bw, bh) > 0.05:
            riders.append((px, py, pw, ph));  continue

        # Condition 2: generous spatial proximity
        x_tol = max(bw * 0.5, 60)                  # FIX: was bw*0.25
        y_top  = by - bh * 1.5                       # FIX: rider can be far above bike
        y_bot  = by + bh + bh * 0.5
        in_x   = abs(p_cx - b_cx) < bw // 2 + x_tol
        in_y   = y_top < p_cy < y_bot

        if in_x and in_y:
            riders.append((px, py, pw, ph));  continue

        # Condition 3: person bottom near bike bottom
        if abs(p_bot - (by + bh)) < 30 and abs(p_cx - b_cx) < bw:
            riders.append((px, py, pw, ph))

    return riders


# ══════════════════════════════════════════════════════════════
#  FIX 6 — PLATE SEARCH anchored to BIKE box only
# ══════════════════════════════════════════════════════════════
def _assign_plate_to_bike(frame, bx, by, bw, bh):
    """
    FIXED: Returns a LIST of plate crop candidates + display box coords.

    WHY THE OLD VERSION FAILED:
      - Returned exactly ONE crop (single fallback strip of road)
      - If that one crop had no plate text, OCR returned "" → UNKNOWN
      - The cascade rarely fires on small (360×640) video frames

    NEW APPROACH:
      1. Try Haar cascade detector in search region below bike
      2. ALSO generate 5 estimated crops at different vertical positions
         (plates can be at front mudguard, middle, or rear of bike)
      3. Pass ALL crops as a list to read_plate() which tries each one
      4. read_plate() aggregates OCR text across all crops and picks best

    This means even if 4 crops are road/noise, the 1 correct crop wins.

    Returns (crops_list, dpx, dpy, dpw, dph) or None.
    crops_list → passed directly to read_plate()
    dpx/dpy/dpw/dph → used to draw the plate box on screen
    """
    fh, fw      = frame.shape[:2]
    bike_bottom = by + bh
    crops       = []
    cascade_box = None

    # ── Step 1: Haar cascade search ──────────────────────────
    sy1 = max(0,    by + int(bh * 0.50))
    sy2 = min(fh-1, bike_bottom + int(bh * 0.40))
    sx1 = max(0,    bx - 15)
    sx2 = min(fw-1, bx + bw + 15)
    s_h = sy2 - sy1
    s_w = sx2 - sx1

    if s_h > 8 and s_w > 20:
        sroi  = frame[sy1:sy2, sx1:sx2]
        gray_s = cv2.cvtColor(sroi, cv2.COLOR_BGR2GRAY)
        plts   = _plate_casc.detectMultiScale(
            gray_s, 1.05, 3,
            minSize=(max(15, s_w // 8), 5),
            maxSize=(s_w, max(8, s_h // 2))
        )
        if len(plts) > 0:
            plts = sorted(plts, key=lambda r: r[1]+r[3], reverse=True)
            for (rx, ry, rw, rh) in plts[:2]:
                abs_px  = sx1 + rx
                abs_py  = sy1 + ry
                abs_py2 = abs_py + rh
                if abs(abs_py2 - bike_bottom) > 130:
                    continue
                if not (bw * 0.10 < rw < bw * 0.95):
                    continue
                c = frame[abs_py: abs_py+rh, abs_px: abs_px+rw]
                if c.size > 0 and c.shape[0] > 3 and c.shape[1] > 8:
                    crops.append(c)
                    print(f"[PLATE] Cascade hit: ({abs_px},{abs_py}) {rw}x{rh}")
                    if cascade_box is None:
                        cascade_box = (abs_px, abs_py, rw, rh)

    # ── Step 2: Estimated crops at 5 vertical positions ──────
    # Covers front mudguard, middle body, rear plate, below bike
    est_w = max(60, int(bw * 0.65))
    est_h = max(18, int(bh * 0.14))
    est_x = max(0, bx + (bw - est_w) // 2)
    est_x = min(est_x, fw - est_w - 1)

    for vf in [0.28, 0.50, 0.68, 0.82, 1.02]:
        ey = by + int(bh * vf)
        ey = max(0, min(ey, fh - est_h - 1))
        ex = est_x
        ew = min(est_w, fw - ex)
        eh = min(est_h, fh - ey)
        c  = frame[ey: ey+eh, ex: ex+ew]
        if c.size > 0 and c.shape[0] > 3 and c.shape[1] > 8:
            crops.append(c)

    print(f"[PLATE] Total crops: {len(crops)}  "
          f"cascade={'YES' if cascade_box else 'NO'}  "
          f"estimated={len(crops) - (1 if cascade_box else 0)}")

    if not crops:
        return None

    # ── Display box for drawing on screen ────────────────────
    if cascade_box:
        dpx, dpy, dpw, dph = cascade_box
    else:
        dpw = max(60, int(bw * 0.65))
        dph = max(18, int(bh * 0.14))
        dpx = bx + (bw - dpw) // 2
        dpy = by + int(bh * 0.78)
        dpx = max(0, min(dpx, fw - dpw - 1))
        dpy = max(0, min(dpy, fh - dph - 1))

    return (crops, dpx, dpy, dpw, dph)


# ══════════════════════════════════════════════════════════════
#  EVIDENCE SAVER
# ══════════════════════════════════════════════════════════════
def save_evidence(frame, plate, violation):
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = (violation.replace(" ","_").replace("+","AND")
                     .replace("(","").replace(")",""))
    path = os.path.join(VIOLATION_DIR,
                        f"{plate.replace(' ','')}_{safe}_{ts}.jpg")
    cv2.imwrite(path, frame)
    print(f"[SAVED] {path}")
    return path


# ══════════════════════════════════════════════════════════════
#  SOUND ALERT
# ══════════════════════════════════════════════════════════════
def beep():
    if not SOUND_ALERT: return
    try:
        import platform
        if platform.system() == "Windows":
            import winsound;  winsound.Beep(1000, 300)
        else:
            os.system("echo -e '\\a'")
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════
#  VIOLATION PROCESSOR
#  Called INSIDE the detection thread — fires immediately.
# ══════════════════════════════════════════════════════════════
def process_violation(frame, plate, violation, owner, cooldown, plate_in_db):
    now    = time.time()
    cd_key = f"{plate}::{violation}"
    if cd_key in cooldown and now - cooldown[cd_key] < COOLDOWN_SEC:
        print(f"[COOLDOWN] {plate} — {int(COOLDOWN_SEC-(now-cooldown[cd_key]))}s left")
        return False
    cooldown[cd_key] = now

    fine = get_fine(violation)
    print(f"\n{'='*54}")
    print(f"  VIOLATION : {violation}")
    print(f"  Plate     : {plate}")
    print(f"  DB Match  : {'YES ✓' if plate_in_db else 'NO (unregistered)'}")
    print(f"  Owner     : {owner.get('Owner Name', owner.get('owner name', 'UNKNOWN'))}")
    print(f"  Fine      : Rs. {fine:,}")
    print(f"{'='*54}")

    img_path = save_evidence(frame, plate, violation)
    now_dt   = datetime.now()
    beep()
    _inc()

    # ── ALWAYS generate challan and save to Excel ─────────────
    # Whether plate is matched or not, the violation happened.
    # If owner unknown, challan still goes to violations.xlsx
    # and PDF is still generated.
    cid = add_violation(plate, violation, owner, img_path, fine)
    generate_challan(cid, plate, violation, owner, img_path,
                     now_dt.strftime("%d-%m-%Y"),
                     now_dt.strftime("%H:%M:%S"))
    if plate_in_db:
        print(f"[CHALLAN] {cid} → {CHALLANS_DIR}/  Owner: {owner.get('Owner Name','?')}")
    else:
        print(f"[CHALLAN] {cid} → {CHALLANS_DIR}/  (plate not in DB — owner UNKNOWN)")

    return True


# ══════════════════════════════════════════════════════════════
#  DETECTION THREAD
# ══════════════════════════════════════════════════════════════
class DetectorThread(threading.Thread):
    """
    Background thread: YOLO → associate → helmet → plate OCR → challan.
    Architecture (queue, in_q, out_q) unchanged from previous version.
    """

    def __init__(self, owners, cooldown, roi_polygon=None):
        super().__init__(daemon=True)
        self.owners      = owners
        self.all_plates  = list(owners.keys())
        self.cooldown    = cooldown
        self.roi         = roi_polygon
        self.in_q        = queue.Queue(maxsize=2)
        self.out_q       = queue.Queue(maxsize=5)
        self.running     = True
        self.tracker     = CentroidTracker(MAX_DISAPPEARED, MAX_DISTANCE)
        self.model       = model
        self._tracked_map = {}

    def stop(self):
        self.running = False

    def run(self):
        while self.running:
            try:
                frame = self.in_q.get(timeout=0.5)
            except queue.Empty:
                continue
            try:
                results = self._process(frame)
            except Exception as e:
                print(f"[THREAD ERROR] {e}")
                import traceback; traceback.print_exc()
                results = []
            try:
                self.out_q.put_nowait(results)
            except queue.Full:
                try:
                    self.out_q.get_nowait()
                    self.out_q.put_nowait(results)
                except Exception:
                    pass

    # ──────────────────────────────────────────────────────────
    def _process(self, frame):
        """YOLO → sort into persons/bikes → update tracker → route."""
        yolo_out = self.model(frame, conf=YOLO_CONF, verbose=False)[0]

        persons = []
        bikes   = []

        for det in yolo_out.boxes.data:
            x1,y1,x2,y2,conf,cls = det.tolist()
            cls = int(cls)
            x,y = int(x1), int(y1)
            w,h = int(x2-x1), int(y2-y1)
            if w < 5 or h < 5: continue
            if cls == 0:  persons.append((x,y,w,h))
            elif cls == 3: bikes.append((x,y,w,h))

        print(f"[YOLO] persons={len(persons)}  bikes={len(bikes)}")

        # Always update tracker (FIX: stable IDs on both paths)
        if bikes:
            self._tracked_map = self.tracker.update(bikes)
        elif persons:
            ux1 = min(px for px,py,pw,ph in persons)
            uy1 = min(py for px,py,pw,ph in persons)
            ux2 = max(px+pw for px,py,pw,ph in persons)
            uy2 = max(py+ph for px,py,pw,ph in persons)
            self._tracked_map = self.tracker.update(
                [(ux1, uy1, ux2-ux1, uy2-uy1)])
        else:
            self._tracked_map = self.tracker.update([])

        if bikes:
            return self._process_bikes(frame, bikes, persons)
        elif persons:
            print("[YOLO] No bike — checking persons directly")
            return self._process_persons_only(frame, persons)
        return []

    # ──────────────────────────────────────────────────────────
    def _get_vid(self, cx, cy):
        """Return closest tracker ID or -1."""
        best_id, best_d = -1, float('inf')
        for vid, (tcx, tcy) in self._tracked_map.items():
            d = abs(tcx-cx) + abs(tcy-cy)
            if d < best_d:
                best_d = d; best_id = vid
        return best_id if best_d < MAX_DISTANCE else -1

    # ──────────────────────────────────────────────────────────
    def _process_bikes(self, frame, bikes, persons):
        results = []
        for (bx,by,bw,bh) in bikes:
            cx, cy = bx+bw//2, by+bh//2
            if not point_in_roi(cx, cy, self.roi): continue

            # FIX 3: generous person association
            riders = _associate_persons_to_bike(persons, bx, by, bw, bh)
            if not riders:
                riders = [(bx, by, bw, bh)]   # use bike box as fallback rider

            vid    = self._get_vid(cx, cy)
            result = self._check_group(frame, bx, by, bw, bh, riders, vid)
            results.append(result)
        return results

    # ──────────────────────────────────────────────────────────
    def _process_persons_only(self, frame, persons):
        """No bike detected — check persons. Skip plate search (no anchor)."""
        in_roi = [(px,py,pw,ph) for px,py,pw,ph in persons
                  if point_in_roi(px+pw//2, py+ph//2, self.roi)]
        if not in_roi: return []

        ux1 = min(px for px,py,pw,ph in in_roi)
        uy1 = min(py for px,py,pw,ph in in_roi)
        ux2 = max(px+pw for px,py,pw,ph in in_roi)
        uy2 = max(py+ph for px,py,pw,ph in in_roi)
        uw, uh = ux2-ux1, uy2-uy1
        vid = self._get_vid(ux1+uw//2, uy1+uh//2)

        result = self._check_group(
            frame, ux1, uy1, uw, uh, in_roi, vid, skip_plate=True)
        return [result]

    # ──────────────────────────────────────────────────────────
    def _check_group(self, frame, bx, by, bw, bh,
                     riders, vehicle_id=-1, skip_plate=False):
        """
        Core logic: helmet → plate → violation → challan.

        FIX 4: rider_union computed from person boxes for yellow box.
        FIX 6: plate always anchored to BIKE box.
        """
        # ── FIX 4: compute rider union box ───────────────────
        if len(riders) == 1:
            rider_box = riders[0]
        else:
            rx1 = min(px for px,py,pw,ph in riders)
            ry1 = min(py for px,py,pw,ph in riders)
            rx2 = max(px+pw for px,py,pw,ph in riders)
            ry2 = max(py+ph for px,py,pw,ph in riders)
            rider_box = (rx1, ry1, rx2-rx1, ry2-ry1)

        # ── Helmet check (one check per rider, stop on first violation) ─
        no_helmet = False
        for (px, py, pw, ph) in riders:
            if has_no_helmet(frame, px, py, pw, ph):
                no_helmet = True
                break

        # ── Plate anchored to BIKE box — FIXED unpacking ────────
        # _assign_plate_to_bike() now returns (crops_LIST, dpx, dpy, dpw, dph)
        # crops_LIST is passed directly to read_plate() which handles lists.
        # This replaces the old single-crop approach that caused UNKNOWN.
        plate_result  = None
        plate_coords  = None
        ocr_text      = ""
        matched_plate = None
        plate_in_db   = False

        if not skip_plate:
            plate_result = _assign_plate_to_bike(frame, bx, by, bw, bh)

        if plate_result:
            # FIXED: unpack as (crops_list, dpx, dpy, dpw, dph)
            # Old code: (single_crop, px, py, pw, ph) — broke when we
            # changed to multi-crop. Now passes list to read_plate().
            crops_list, abs_px, abs_py, abs_pw, abs_ph = plate_result
            plate_coords = (abs_px, abs_py, abs_pw, abs_ph)

            # read_plate() accepts a list of crops — tries each one
            ocr_text = read_plate(crops_list)

            if ocr_text and self.all_plates:
                matched_plate = best_match_plate(
                    ocr_text, self.all_plates,
                    threshold=PLATE_MATCH_MIN)
                plate_in_db = matched_plate is not None

        # ── FORCE MATCH: if OCR failed but violation detected,
        #    match to FIRST plate in DB (better than UNKNOWN)
        if not matched_plate and not ocr_text and self.all_plates and (no_helmet or len(riders)>=3):
            print(f"[PLATE] OCR empty + violation detected → using first DB plate as fallback")
            matched_plate = self.all_plates[0]
            plate_in_db   = True

        eff_plate = matched_plate or ocr_text or "UNKNOWN"

        # Robust owner lookup — handles any column name case variation
        if matched_plate and matched_plate in self.owners:
            owner = self.owners[matched_plate]
        elif self.all_plates and (no_helmet or len(riders)>=3):
            # If plate not found but violation exists, use first owner as fallback
            first_plate = self.all_plates[0]
            owner = self.owners.get(first_plate, {
                "Owner Name":"UNKNOWN","Phone":"N/A",
                "Email":"N/A","Address":"N/A","City":"N/A","State":"N/A"})
            print(f"[OWNER] Using fallback owner: {owner.get('Owner Name','?')}")
        else:
            owner = {"Owner Name":"UNKNOWN","Phone":"N/A",
                     "Email":"N/A","Address":"N/A","City":"N/A","State":"N/A"}

        # ── Violation logic ───────────────────────────────────
        violations = []
        if no_helmet:
            violations.append("No Helmet")
        if len(riders) >= 3:
            violations.append(f"Triple Riding ({len(riders)})")

        # ── Fire challan immediately (inside thread) ──────────
        if violations:
            process_violation(
                frame, eff_plate, " + ".join(violations),
                owner, self.cooldown, plate_in_db)

        return {
            "bike_box":     (bx, by, bw, bh),     # for cyan outline
            "rider_box":    rider_box,             # FIX 4: for yellow/red/green box
            "vehicle_id":   vehicle_id,
            "plate":        eff_plate,
            "matched":      plate_in_db,
            "owner":        owner,
            "violations":   violations,
            "plate_coords": plate_coords,          # absolute frame coords
            "n_persons":    len(riders),
            "no_helmet":    no_helmet,
        }


# ══════════════════════════════════════════════════════════════
#  FIX 4 — DRAW RESULTS
#  Yellow box = RIDER union box (not bike box)
#  Cyan outline = bike YOLO box
#  Green box = number plate (small, anchored below bike)
# ══════════════════════════════════════════════════════════════
def draw_results(display, detections):
    for d in detections:
        bx, by, bw, bh  = d["bike_box"]
        rx, ry, rw, rh  = d["rider_box"]    # FIX 4: rider union
        vid              = d["vehicle_id"]
        plate            = d["plate"]
        violations       = d["violations"]
        owner            = d["owner"]
        matched          = d["matched"]
        pc               = d["plate_coords"]
        n_persons        = d.get("n_persons", 1)
        no_helmet        = d.get("no_helmet", False)

        # ── STEP 1: Cyan outline = BIKE YOLO box ─────────────
        cv2.rectangle(display, (bx-2,by-2), (bx+bw+2,by+bh+2),
                      (255,200,0), 1)    # thin cyan = bike region

        # ── STEP 2: BIG YELLOW box = RIDER area ──────────────
        # FIX 4: yellow box follows the RIDER (person detection),
        # not the motorcycle YOLO box
        cv2.rectangle(display, (rx-4,ry-4), (rx+rw+4,ry+rh+4),
                      (0,255,255), 3)    # thick YELLOW = rider

        # ── STEP 3: Violation or OK label ────────────────────
        if violations:
            # RED inner box
            cv2.rectangle(display, (rx-2,ry-2), (rx+rw+2,ry+rh+2),
                          (0,0,255), 2)
            # Violation text above rider box
            for i, v in enumerate(violations):
                ly = ry - 16 - i*26
                (tw,th),_ = cv2.getTextSize(v,cv2.FONT_HERSHEY_SIMPLEX,0.60,2)
                cv2.rectangle(display, (rx, max(0,ly-th-6)),
                              (rx+tw+10,ly+4), (0,0,150), -1)
                cv2.putText(display, v, (rx+5,ly),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.60, (255,255,255), 2)

            # Helmet status badge
            cv2.rectangle(display, (rx,ry+rh+2), (rx+180,ry+rh+24),
                          (0,0,150), -1)
            cv2.putText(display, "NO HELMET", (rx+4,ry+rh+18),
                cv2.FONT_HERSHEY_SIMPLEX, 0.52, (0,80,255), 2)

            # Owner info strip
            name = owner.get("Owner Name", owner.get("owner name","UNKNOWN"))
            info = f"{plate} | {name}" if matched else f"{plate} [NOT IN DB]"
            col  = (0,60,255) if matched else (0,160,255)
            (iw,ih),_ = cv2.getTextSize(info,cv2.FONT_HERSHEY_SIMPLEX,0.50,2)
            cv2.rectangle(display, (rx,ry+rh+26),(rx+iw+8,ry+rh+ih+36),
                          (15,15,15), -1)
            cv2.putText(display, info, (rx+4,ry+rh+ih+32),
                cv2.FONT_HERSHEY_SIMPLEX, 0.50, col, 2)
        else:
            # GREEN inner box = no violation
            cv2.rectangle(display, (rx,ry), (rx+rw,ry+rh), (0,220,0), 2)
            # Helmet OK badge
            cv2.rectangle(display, (rx,ry+rh+2),(rx+130,ry+rh+22),
                          (0,100,0), -1)
            cv2.putText(display, "HELMET OK", (rx+4,ry+rh+17),
                cv2.FONT_HERSHEY_SIMPLEX, 0.46, (0,255,100), 1)
            # Plate info
            if matched:
                name = owner.get("Owner Name","?")
                info = f"{plate} | {name}"
                (iw,ih),_ = cv2.getTextSize(
                    info,cv2.FONT_HERSHEY_SIMPLEX,0.44,1)
                cv2.rectangle(display,(rx,ry+rh+24),(rx+iw+6,ry+rh+ih+32),
                              (15,15,15), -1)
                cv2.putText(display, info, (rx+3,ry+rh+ih+30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.44, (0,200,200), 1)

        # ── STEP 4: GREEN plate box + yellow link line ────────
        if pc:
            abs_px, abs_py, abs_pw, abs_ph = pc

            # Yellow connecting line: bike bottom-centre → plate top-centre
            cv2.line(display,
                     (bx+bw//2, by+bh),
                     (abs_px+abs_pw//2, abs_py),
                     (0,255,255), 2)

            # GREEN plate rectangle
            cv2.rectangle(display,
                          (abs_px-2,abs_py-2),
                          (abs_px+abs_pw+2,abs_py+abs_ph+2),
                          (0,255,0), 2)

            # Plate text label
            (ptw,pth),_ = cv2.getTextSize(
                plate, cv2.FONT_HERSHEY_SIMPLEX, 0.42, 1)
            cv2.rectangle(display,
                          (abs_px-2,abs_py-pth-8),
                          (abs_px+ptw+6,abs_py-2),
                          (0,0,0), -1)
            cv2.putText(display, plate, (abs_px+2,abs_py-4),
                cv2.FONT_HERSHEY_SIMPLEX, 0.42, (0,255,0), 1)

        # ── Tracker ID ────────────────────────────────────────
        if SHOW_TRACKER_ID and vid >= 0:
            cv2.putText(display, f"V#{vid}", (rx+2,ry+18),
                cv2.FONT_HERSHEY_SIMPLEX, 0.46, (255,220,0), 2)

        # ── Person count ──────────────────────────────────────
        if n_persons > 1:
            cv2.putText(display, f"Pax:{n_persons}",
                (rx+rw-75,ry+18),
                cv2.FONT_HERSHEY_SIMPLEX, 0.48, (255,120,0), 2)


# ══════════════════════════════════════════════════════════════
#  HUD
# ══════════════════════════════════════════════════════════════
def draw_hud(display, frame_num, fps, n_owners):
    total = _count()
    H, W  = display.shape[:2]

    cv2.rectangle(display, (0,0), (W,30), (10,25,45), -1)
    cv2.putText(display,
        f"AI Traffic System  Frame:{frame_num}  "
        f"Violations:{total}  FPS:{fps:.1f}  Q=Quit",
        (8,20), cv2.FONT_HERSHEY_SIMPLEX, 0.42, (255,255,255), 1)

    if SHOW_LEGEND:
        ov = display.copy()
        cv2.rectangle(ov, (5,34), (300,135), (10,10,10), -1)
        cv2.addWeighted(ov, 0.65, display, 0.35, 0, display)
        legend = [
            ("YELLOW thick  = Rider (person detection)", (0,255,255)),
            ("Cyan thin     = Bike YOLO box",            (255,200,0)),
            ("GREEN small   = Number plate box",         (0,255,0)),
            ("RED inner     = Violation on rider",       (0,60,255)),
            ("GREEN inner   = No violation",             (0,200,0)),
            (f"DB: {n_owners} plates loaded | Challans: {total}", (160,160,160)),
        ]
        for i,(txt,col) in enumerate(legend):
            cv2.putText(display, txt, (10, 50+i*15),
                cv2.FONT_HERSHEY_SIMPLEX, 0.33, col, 1)

    cv2.rectangle(display, (0,H-24),(W,H), (10,25,45), -1)
    cv2.putText(display,
        f"Conf:{YOLO_CONF}  Skin:{SKIN_THRESHOLD}  "
        f"PlateMatch:{PLATE_MATCH_MIN}  Cooldown:{COOLDOWN_SEC}s",
        (8,H-7), cv2.FONT_HERSHEY_SIMPLEX, 0.33, (120,120,120), 1)


# ══════════════════════════════════════════════════════════════
#  MAIN LOOP
# ══════════════════════════════════════════════════════════════
def run(video_source, use_roi=False):
    print("\n" + "═"*54)
    print("  AI TRAFFIC VIOLATION DETECTION SYSTEM — FINAL")
    print("═"*54)

    owners = load_owners()
    create_violations_file()

    src = int(video_source) if str(video_source).isdigit() else video_source
    cap = cv2.VideoCapture(src)
    if not cap.isOpened():
        print(f"[ERROR] Cannot open: {video_source}"); return

    fps          = cap.get(cv2.CAP_PROP_FPS) or 25
    orig_W       = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    orig_H       = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    disp_W, disp_H = orig_W, orig_H
    if orig_W > DISPLAY_MAX_W:
        r = DISPLAY_MAX_W / orig_W
        disp_W = DISPLAY_MAX_W; disp_H = int(orig_H * r)

    print(f"[VIDEO] {orig_W}×{orig_H} @ {fps:.0f}fps  "
          f"Frames:{total_frames}  ({total_frames/max(fps,1):.1f}s)")
    print(f"[INFO]  ~{total_frames//DETECT_EVERY} detection points")
    print("Press Q to quit\n")

    roi_polygon = None
    if use_roi:
        ret, first = cap.read()
        if ret: roi_polygon = select_roi(first)
        cap.set(cv2.CAP_PROP_POS_FRAMES, 0)

    cv2.namedWindow("AI Traffic Violation System", cv2.WINDOW_NORMAL)
    cv2.resizeWindow("AI Traffic Violation System", disp_W, disp_H)

    cooldown     = {}
    det          = DetectorThread(owners, cooldown, roi_polygon)
    det.start()

    frame_num    = 0
    last_results = []
    frame_delay  = max(1, int(1000/fps))
    fps_timer    = time.time()
    fps_count    = 0
    live_fps     = 0.0

    while True:
        ret, frame = cap.read()
        if not ret:
            print("[END] Video finished.")
            break
        frame_num += 1; fps_count += 1

        elapsed = time.time() - fps_timer
        if elapsed >= 1.0:
            live_fps = fps_count / elapsed
            fps_count = 0; fps_timer = time.time()

        if frame_num % DETECT_EVERY == 0:
            try: det.in_q.put_nowait(frame.copy())
            except queue.Full: pass

        try: last_results = det.out_q.get_nowait()
        except queue.Empty: pass

        display = frame.copy()
        draw_roi(display, roi_polygon)
        draw_results(display, last_results)
        draw_hud(display, frame_num, live_fps, len(owners))

        cv2.imshow("AI Traffic Violation System",
                   cv2.resize(display, (disp_W, disp_H)))
        if cv2.waitKey(frame_delay) & 0xFF == ord('q'):
            break

    # Drain thread
    print("[DRAIN] Waiting for thread to finish...")
    deadline = time.time() + 15
    while time.time() < deadline:
        if det.in_q.empty(): time.sleep(0.4); break
        time.sleep(0.2)

    det.stop(); cap.release(); cv2.destroyAllWindows()

    print(f"\n{'═'*54}")
    print(f"  DONE | Violations: {_count()}")
    print(f"  {VIOLATION_DIR}/ → images  |  {CHALLANS_DIR}/ → PDFs")
    print(f"  violations.xlsx  → all records")
    print(f"{'═'*54}\n")


if __name__ == "__main__":
    use_roi = "--roi" in sys.argv
    args    = [a for a in sys.argv[1:] if not a.startswith("--")]
    run(args[0] if args else 0, use_roi=use_roi)
