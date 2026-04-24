"""
create_owners.py
Run this ONCE to create bike_owners.xlsx
Add your own bike data here
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Bike Owners"

headers = ["Number Plate", "Owner Name", "Phone", "Email", "Address", "City", "State", "Pincode"]

# Header style
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center")

# -------------------------------------------------------
# YOUR REAL BIKE DATA FROM THE VIDEOS
# -------------------------------------------------------
data = [
    # Video 1 — Triple riding
    ["BR02BS9361", "Akshay Sonawane", "1234567891",
     "akshay.sonawane@gmail.com", "Near Station Road", "Patna", "Bihar", "800001"],

    # Video 2 — No helmet
    ["DL9SCD5588", "Lalit Wagh", "9123456879",
     "lalit.wagh@gmail.com", "Sector 12 Dwarka", "Delhi", "Delhi", "110075"],

    # Extra demo entries
    ["MH12AB1234", "Rahul Sharma",  "9876543210",
     "rahul.sharma@gmail.com",  "123 MG Road",      "Pune",    "Maharashtra", "411001"],
    ["MH14CD5678", "Priya Patil",   "9823456789",
     "priya.patil@gmail.com",   "45 Shivaji Nagar", "Pune",    "Maharashtra", "411005"],
]

for row in data:
    ws.append(row)

# Column widths
widths = [16, 20, 14, 32, 28, 15, 15, 10]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(1, col).column_letter].width = w

wb.save("bike_owners.xlsx")
print("bike_owners.xlsx created!")
for d in data:
    print(f"  {d[0]} → {d[1]}")
