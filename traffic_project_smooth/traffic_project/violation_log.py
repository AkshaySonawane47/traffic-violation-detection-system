"""
violation_log.py
Saves every violation to violations.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

FILE = "violations.xlsx"


def create_violations_file():
    if os.path.exists(FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Violations"

    headers = [
        "Challan ID", "Date", "Time",
        "Number Plate", "Violation Type",
        "Owner Name", "Phone", "Email", "Address",
        "Fine (Rs)", "Image Path", "Status"
    ]
    fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    widths = [14,12,10,16,28,20,14,30,28,12,40,12]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1,c).column_letter].width = w

    wb.save(FILE)
    print(f"[LOG] Created {FILE}")


def add_violation(plate, violation, owner, img_path, fine):
    create_violations_file()
    wb = openpyxl.load_workbook(FILE)
    ws = wb.active

    # Challan ID = EC + date + row number
    row_num = ws.max_row         # existing rows including header
    challan_id = f"EC{datetime.now().strftime('%Y%m%d')}{row_num:04d}"
    now = datetime.now()

    row = [
        challan_id,
        now.strftime("%d-%m-%Y"),
        now.strftime("%H:%M:%S"),
        plate,
        violation,
        owner.get("Owner Name", "UNKNOWN"),
        owner.get("Phone",      "N/A"),
        owner.get("Email",      "N/A"),
        str(owner.get("Address","N/A")) + ", " +
        str(owner.get("City","")) + ", " +
        str(owner.get("State","")),
        fine,
        img_path,
        "Pending"
    ]
    ws.append(row)

    # Row color by violation type
    colors = {
        "No Helmet":    "FCE4D6",
        "Triple":       "DDEBF7",
        "Overspeed":    "FFE0E0",
    }
    color = "FFF2CC"
    for key, c in colors.items():
        if key in violation:
            color = c
            break

    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    new_row = ws.max_row
    for c in range(1, 13):
        ws.cell(new_row, c).fill = fill

    wb.save(FILE)
    print(f"[LOG] Saved → {challan_id} | {plate} | {violation}")
    return challan_id


def get_all_violations():
    create_violations_file()
    wb = openpyxl.load_workbook(FILE)
    ws = wb.active
    hdrs = [ws.cell(1,c).value for c in range(1,13)]
    rows = []
    for r in range(2, ws.max_row+1):
        if ws.cell(r,1).value:
            rows.append({h: ws.cell(r,c).value for c,h in enumerate(hdrs,1)})
    return rows
