"""
create_owners.py
================
Run once to create / reset bike_owners.xlsx.
Add your own entries to the `data` list below.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HEADERS = ["Number Plate","Owner Name","Phone","Email",
           "Address","City","State","Pincode"]
WIDTHS  = [16, 22, 14, 32, 30, 15, 15, 10]

# ─── YOUR REGISTERED VEHICLES ───────────────────────────────
data = [
    # Plate           Name               Phone          Email
    # Address                    City       State       Pincode
    ["BR02BS9361", "Akshay Sonawane", "9234567891",
     "akshay.sonawane@gmail.com", "Near Station Road", "Patna", "Bihar", "800001"],

    ["DL9SCD5588", "Lalit Wagh",      "9123456879",
     "lalit.wagh@gmail.com", "Sector 12 Dwarka", "Delhi", "Delhi", "110075"],

    ["MH12AB1234", "Rahul Sharma",    "9876543210",
     "rahul.sharma@gmail.com", "123 MG Road", "Pune", "Maharashtra", "411001"],

    ["MH14CD5678", "Priya Patil",     "9823456789",
     "priya.patil@gmail.com", "45 Shivaji Nagar", "Pune", "Maharashtra", "411005"],

    ["KA03MN7890", "Suresh Kumar",    "9765432100",
     "suresh.kumar@gmail.com", "12 Brigade Road", "Bengaluru", "Karnataka", "560001"],
]
# ─────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Bike Owners"

# Header row
hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hdr_font = Font(color="FFFFFF", bold=True, size=10)
for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
    cell = ws.cell(1, col, h)
    cell.fill      = hdr_fill
    cell.font      = hdr_font
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[cell.column_letter].width = w

ws.row_dimensions[1].height = 20

# Data rows — alternate background
for i, row in enumerate(data):
    ws.append(row)
    fill_col = "EBF5FB" if i % 2 == 0 else "FFFFFF"
    for c in range(1, len(HEADERS)+1):
        ws.cell(i+2, c).fill = PatternFill(
            start_color=fill_col, end_color=fill_col, fill_type="solid")

wb.save("bike_owners.xlsx")
print("bike_owners.xlsx created successfully!")
for d in data:
    print(f"  {d[0]:<14}  →  {d[1]}")
