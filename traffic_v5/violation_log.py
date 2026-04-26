"""
violation_log.py
================
Saves every detected violation to violations.xlsx.
Each row = one challan record with colour coding by violation type.

Functions
---------
create_violations_file()  — creates file with header if not exists
add_violation(...)        — appends a new violation row, returns challan_id
get_all_violations()      — returns all rows as list of dicts
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

try:
    from config.settings import VIOLATIONS_FILE
except ImportError:
    VIOLATIONS_FILE = "violations.xlsx"

FILE = VIOLATIONS_FILE

# Column headers + widths
HEADERS = [
    "Challan ID", "Date", "Time",
    "Number Plate", "Violation Type",
    "Owner Name", "Phone", "Email", "Address",
    "Fine (Rs)", "Image Path", "Status"
]
WIDTHS = [14, 12, 10, 16, 30, 22, 14, 30, 35, 12, 40, 12]

# Violation-type colour coding
VIOL_COLORS = {
    "No Helmet":    "FCE4D6",   # light orange
    "Triple":       "DDEBF7",   # light blue
    "Overspeed":    "FFE0E0",   # light red
}
DEFAULT_COLOR = "FFF2CC"        # light yellow


def _thin_border():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def create_violations_file():
    """Create violations.xlsx with styled header row if it doesn't exist."""
    if os.path.exists(FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Violations"

    # Header row
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    for c, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(1, c, h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22
    wb.save(FILE)
    print(f"[LOG] Created {FILE}")


def add_violation(plate, violation, owner, img_path, fine):
    """
    Append a violation row to violations.xlsx.

    Parameters
    ----------
    plate     : str  — number plate (e.g. 'MH12AB1234')
    violation : str  — violation description
    owner     : dict — owner info from bike_owners.xlsx
    img_path  : str  — path to evidence image
    fine      : int  — fine amount in Rs.

    Returns
    -------
    challan_id : str  (e.g. 'EC202504250003')
    """
    create_violations_file()
    wb  = openpyxl.load_workbook(FILE)
    ws  = wb.active
    now = datetime.now()

    # Challan ID: EC + YYYYMMDD + 4-digit row number
    row_num    = ws.max_row   # includes header
    challan_id = f"EC{now.strftime('%Y%m%d')}{row_num:04d}"

    address = (str(owner.get("Address", "N/A")) + ", " +
               str(owner.get("City", "")) + ", " +
               str(owner.get("State", "")))

    row_data = [
        challan_id,
        now.strftime("%d-%m-%Y"),
        now.strftime("%H:%M:%S"),
        plate,
        violation,
        owner.get("Owner Name", "UNKNOWN"),
        owner.get("Phone",      "N/A"),
        owner.get("Email",      "N/A"),
        address,
        fine,
        img_path,
        "Pending"
    ]
    ws.append(row_data)

    # Colour-code by violation type
    color = DEFAULT_COLOR
    for key, c in VIOL_COLORS.items():
        if key in violation:
            color = c
            break

    fill   = PatternFill(start_color=color, end_color=color, fill_type="solid")
    border = _thin_border()
    new_r  = ws.max_row
    for c in range(1, len(HEADERS) + 1):
        cell        = ws.cell(new_r, c)
        cell.fill   = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center")

    wb.save(FILE)
    print(f"[LOG] {challan_id} | {plate} | {violation}")
    return challan_id


def get_all_violations():
    """Return all violation rows as a list of dicts."""
    create_violations_file()
    wb   = openpyxl.load_workbook(FILE)
    ws   = wb.active
    hdrs = [ws.cell(1, c).value for c in range(1, len(HEADERS) + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value:
            rows.append({h: ws.cell(r, c).value
                         for c, h in enumerate(hdrs, 1)})
    return rows
