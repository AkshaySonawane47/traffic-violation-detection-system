"""
detector.py  -  Traffic Violation Detection System
===================================================
FIXES IN THIS VERSION:
  - Video plays smoothly (no lag)
  - Detection runs in background thread (does NOT block video)
  - OCR runs only every 15 frames (not every frame)
  - Resize frame before processing (faster on 4GB RAM)
  - Plate matched first -> then check violation
  - Helmet: face visible = NO helmet, face hidden = helmet OK
  - Only bikes in Excel database get challan

USAGE:
  python detector.py video1_triple.mp4
  python detector.py video2_helmet.mp4
  python detector.py          <- webcam
"""

import cv2
import os
import sys
import time
import threading
import queue
import numpy as np
from datetime import datetime
import openpyxl

from violation_log import create_violations_file, add_violation
from challan_pdf   import generate_challan, get_fine
from plate_ocr     import read_plate, best_match_plate

# ═══════════════════════════════════════════════
#  SETTINGS  (tune these for your PC)
# ═══════════════════════════════════════════════
DETECT_EVERY     = 15     # run detection every N frames (video plays every frame)
COOLDOWN_SEC     = 30     # seconds before same plate flagged again
PLATE_MATCH_MIN  = 0.55   # 55% similarity needed to match plate
FACE_MIN_NBRS    = 5      # face cascade strictness (higher = fewer false alarms)
TRIPLE_MIN       = 3      # persons count for triple riding
RESIZE_WIDTH     = 480    # resize frame before detection (smaller = faster)
OWNERS_FILE      = "bike_owners.xlsx"
VIOLATION_DIR    = "violation_images"
os.makedirs(VIOLATION_DIR, exist_ok=True)


# ═══════════════════════════════════════════════
#  LOAD OWNERS FROM EXCEL
# ═══════════════════════════════════════════════
def load_owners():
    owners = {}
    if not os.path.exists(OWNERS_FILE):
        print(f"[ERROR] {OWNERS_FILE} not found. Run create_owners.py first.")
        return owners
    wb   = openpyxl.load_workbook(OWNERS_FILE)
    ws   = wb.active
    hdrs = [ws.cell(1, c).value for c in range(1, 9)]
    for r in range(2, ws.max_row + 1):
        plate = ws.cell(r, 1).value
        if plate:
            key = str(plate).strip().upper().replace(" ", "")
            owners[key] = {h: ws.cell(r, c).value for c, h in enumerate(hdrs, 1)}
    print(f"\n[DB] {len(owners)} owners loaded:")
    for p, o in owners.items():
        print(f"     {p}  ->  {o.get('Owner Name','?')}")
    return owners


# ═══════════════════════════════════════════════
#  CASCADE CLASSIFIERS  (loaded once)
# ═══════════════════════════════════════════════
_face_casc  = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_alt2.xml")
_upper_casc = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_upperbody.xml")
_plate_casc = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_russian_plate_number.xml")


# ═══════════════════════════════════════════════
#  HELMET CHECK
# ═══════════════════════════════════════════════
def check_helmet(frame, x, y, w, h):
    """
    True  = helmet present (OK)
    False = no helmet (VIOLATION)

    If face is clearly visible -> rider not wearing helmet
    If face hidden/covered    -> helmet is on
    """
    head_end = y + int(h * 0.42)
    roi = frame[max(0, y): head_end, max(0, x): x + w]
    if roi.size == 0:
        return True

    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    gray = cv2.equalizeHist(gray)

    faces = _face_casc.detectMultiScale(
        gray,
        scaleFactor  = 1.08,
        minNeighbors = FACE_MIN_NBRS,
        minSize      = (18, 18)
    )
    if len(faces) > 0:
        return False   # face visible = no helmet
    return True        # face hidden = helmet on


# ═══════════════════════════════════════════════
#  PERSON COUNT
# ═══════════════════════════════════════════════
def count_persons(frame, x, y, w, h):
    pad  = 10
    roi  = frame[max(0, y-pad): y+h+pad, max(0, x-pad): x+w+pad]
    if roi.size == 0:
        return 1
    gray  = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    upper = _upper_casc.detectMultiScale(gray, 1.1, 3, minSize=(20, 20))
    return max(1, len(upper))


# ═══════════════════════════════════════════════
#  FIND PLATE CROP
# ═══════════════════════════════════════════════
def find_plate(frame, x, y, w, h):
    """Returns (crop, px, py, pw, ph) or None"""
    # Search bottom 45% of bounding box
    sy   = y + int(h * 0.55)
    sroi = frame[sy: y+h, max(0,x): x+w]
    if sroi.size > 0:
        gray  = cv2.cvtColor(sroi, cv2.COLOR_BGR2GRAY)
        plts  = _plate_casc.detectMultiScale(gray, 1.05, 4, minSize=(20, 8))
        if len(plts) > 0:
            px2, py2, pw2, ph2 = plts[0]
            crop = frame[sy+py2: sy+py2+ph2, x+px2: x+px2+pw2]
            return (crop, x+px2, sy+py2, pw2, ph2)

    # Fallback: estimated plate position
    ew   = int(w * 0.55)
    eh   = max(12, int(h * 0.10))
    ex   = x + (w - ew) // 2
    ey   = y + int(h * 0.78)
    crop = frame[max(0,ey): ey+eh, max(0,ex): ex+ew]
    if crop.size > 0:
        return (crop, ex, ey, ew, eh)
    return None


# ═══════════════════════════════════════════════
#  SAVE IMAGE
# ═══════════════════════════════════════════════
def save_evidence(frame, plate, violation):
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = violation.replace(" ", "_").replace("+", "AND")
    path = os.path.join(VIOLATION_DIR, f"{plate}_{safe}_{ts}.jpg")
    cv2.imwrite(path, frame)
    return path


# ═══════════════════════════════════════════════
#  PROCESS VIOLATION
# ═══════════════════════════════════════════════
def process_violation(frame, plate, violation, owner, cooldown):
    now = time.time()
    if plate in cooldown and now - cooldown[plate] < COOLDOWN_SEC:
        return False
    cooldown[plate] = now

    fine = get_fine(violation)
    print(f"\n{'='*50}")
    print(f"  VIOLATION : {violation}")
    print(f"  Plate     : {plate}")
    print(f"  Owner     : {owner.get('Owner Name','UNKNOWN')}")
    print(f"  Fine      : Rs.{fine}")
    print(f"{'='*50}")

    img_path   = save_evidence(frame, plate, violation)
    now_dt     = datetime.now()
    challan_id = add_violation(plate, violation, owner, img_path, fine)
    generate_challan(
        challan_id, plate, violation, owner, img_path,
        now_dt.strftime("%d-%m-%Y"),
        now_dt.strftime("%H:%M:%S")
    )
    print(f"[CHALLAN] {challan_id} saved -> challans/")
    return True


# ═══════════════════════════════════════════════
#  DETECTION WORKER  (runs in background thread)
# ═══════════════════════════════════════════════
class DetectionWorker(threading.Thread):
    """
    Runs in background thread so video plays smoothly.
    Main thread just shows video.
    This thread does detection work.
    """
    def __init__(self, owners, cooldown):
        super().__init__(daemon=True)
        self.owners       = owners
        self.all_plates   = list(owners.keys())
        self.cooldown     = cooldown
        self.input_q      = queue.Queue(maxsize=2)   # holds frames to process
        self.result_q     = queue.Queue(maxsize=5)   # holds detection results
        self.running      = True

    def stop(self):
        self.running = False

    def run(self):
        """Background loop: get frame -> detect -> put result"""
        while self.running:
            try:
                frame, frame_num = self.input_q.get(timeout=0.5)
            except queue.Empty:
                continue

            # Resize for faster processing
            scale  = RESIZE_WIDTH / max(frame.shape[1], 1)
            small  = cv2.resize(frame, (RESIZE_WIDTH, int(frame.shape[0] * scale)))

            # Background subtraction on small frame
            detections = self._detect(frame, small, scale)

            # Put results back (non-blocking)
            try:
                self.result_q.put_nowait(detections)
            except queue.Full:
                pass   # drop if main thread hasn't consumed yet

    def _detect(self, orig_frame, small_frame, scale):
        """
        Core detection on one frame.
        Returns list of detection dicts to draw on screen.
        """
        results = []

        # Motion mask
        if not hasattr(self, '_bg'):
            self._bg = cv2.createBackgroundSubtractorMOG2(
                history=150, varThreshold=40, detectShadows=False)

        mask = self._bg.apply(small_frame)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE,
                    cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7,7)))
        mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN,
                    cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5,5)))

        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        for cnt in contours:
            area = cv2.contourArea(cnt)
            # Bike-sized area on small frame
            if area < 800 or area > 50000:
                continue

            sx, sy, sw, sh = cv2.boundingRect(cnt)
            if sy < 15:
                continue
            aspect = sw / max(sh, 1)
            if aspect < 0.4 or aspect > 6.0:
                continue

            # Scale coords back to original frame size
            x  = int(sx / scale)
            y  = int(sy / scale)
            w  = int(sw / scale)
            h  = int(sh / scale)

            # Clamp to frame bounds
            fh, fw = orig_frame.shape[:2]
            x  = max(0, min(x, fw-1))
            y  = max(0, min(y, fh-1))
            w  = min(w, fw - x)
            h  = min(h, fh - y)
            if w < 20 or h < 20:
                continue

            # --- STEP 1: Read plate ---
            plate_data    = find_plate(orig_frame, x, y, w, h)
            matched_plate = None
            plate_coords  = None
            ocr_raw       = ""

            if plate_data:
                crop, px, py, pw, ph = plate_data
                plate_coords = (px, py, pw, ph)
                ocr_raw      = read_plate(crop)
                if ocr_raw:
                    matched_plate = best_match_plate(
                        ocr_raw, self.all_plates,
                        threshold=PLATE_MATCH_MIN
                    )

            # --- STEP 2: Only check violation if plate matched ---
            if not matched_plate:
                results.append({
                    "box":    (x, y, w, h),
                    "plate":  None,
                    "ocr":    ocr_raw,
                    "owner":  None,
                    "violations": [],
                    "plate_coords": plate_coords,
                })
                continue

            owner = self.owners[matched_plate]

            # --- STEP 3: Check violations ---
            violations = []

            if not check_helmet(orig_frame, x, y, w, h):
                violations.append("No Helmet")

            n = count_persons(orig_frame, x, y, w, h)
            if n >= TRIPLE_MIN:
                violations.append(f"Triple Riding ({n} persons)")

            # --- STEP 4: Process if violation found ---
            if violations:
                vstr = " + ".join(violations)
                process_violation(
                    orig_frame, matched_plate, vstr, owner, self.cooldown
                )

            results.append({
                "box":         (x, y, w, h),
                "plate":       matched_plate,
                "ocr":         ocr_raw,
                "owner":       owner,
                "violations":  violations,
                "plate_coords": plate_coords,
            })

        return results


# ═══════════════════════════════════════════════
#  DRAW RESULTS ON DISPLAY FRAME
# ═══════════════════════════════════════════════
def draw_results(display, detections, total_viol):
    for d in detections:
        x, y, w, h  = d["box"]
        plate        = d["plate"]
        violations   = d["violations"]
        owner        = d["owner"]
        plate_coords = d["plate_coords"]
        ocr_raw      = d["ocr"]

        if plate is None:
            # No plate match - draw faint grey box only
            cv2.rectangle(display, (x,y), (x+w,y+h), (60,60,60), 1)
            if ocr_raw:
                cv2.putText(display, f"?{ocr_raw}",
                    (x, y-4), cv2.FONT_HERSHEY_SIMPLEX, 0.38, (60,60,180), 1)
            continue

        owner_name = owner.get("Owner Name","UNKNOWN") if owner else "UNKNOWN"

        if violations:
            # RED - violation found
            cv2.rectangle(display, (x-4,y-4), (x+w+4,y+h+4), (0,0,255), 3)

            # Violation labels
            for i, v in enumerate(violations):
                label_y = y - 8 - i*22
                (tw, th), _ = cv2.getTextSize(v, cv2.FONT_HERSHEY_SIMPLEX, 0.56, 2)
                cv2.rectangle(display, (x, label_y-th-4), (x+tw+6, label_y+2), (0,0,220), -1)
                cv2.putText(display, v, (x+3, label_y),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.56, (255,255,255), 2)

            # Owner + plate below box
            info = f"{plate} | {owner_name}"
            cv2.putText(display, info, (x, y+h+18),
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0,80,255), 2)

        else:
            # GREEN - plate matched, no violation
            cv2.rectangle(display, (x,y), (x+w,y+h), (0,200,0), 2)
            cv2.putText(display, f"{plate} OK",
                (x, y+h+18), cv2.FONT_HERSHEY_SIMPLEX, 0.48, (0,200,0), 2)

        # Plate region box
        if plate_coords:
            px, py, pw, ph = plate_coords
            cv2.rectangle(display, (px,py), (px+pw,py+ph), (0,220,220), 1)
            cv2.putText(display, plate,
                (px, py-4), cv2.FONT_HERSHEY_SIMPLEX, 0.42, (0,220,220), 1)


# ═══════════════════════════════════════════════
#  MAIN  -  smooth video playback loop
# ═══════════════════════════════════════════════
def run(video_source):
    print("\n" + "="*50)
    print("  TRAFFIC VIOLATION DETECTION SYSTEM")
    print("  Video plays smooth, detection in background")
    print("="*50)

    owners = load_owners()
    create_violations_file()

    if not owners:
        print("[ERROR] No owners loaded. Cannot run.")
        return

    cap = cv2.VideoCapture(video_source)
    if not cap.isOpened():
        print(f"[ERROR] Cannot open: {video_source}")
        return

    fps = cap.get(cv2.CAP_PROP_FPS) or 25
    W   = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    H   = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    print(f"[VIDEO] {W}x{H} @ {fps:.0f}fps")
    print("[INFO]  Video plays every frame. Detection every 15 frames.")
    print("Press Q to quit\n")

    cooldown   = {}
    worker     = DetectionWorker(owners, cooldown)
    worker.start()

    frame_num   = 0
    total_viol  = 0
    last_results = []     # last detection results to draw
    frame_delay  = max(1, int(1000 / fps))  # ms per frame for cv2.waitKey

    while True:
        ret, frame = cap.read()
        if not ret:
            print("[END] Video finished.")
            break
        frame_num += 1

        # Send frame to background worker every N frames
        if frame_num % DETECT_EVERY == 0:
            try:
                worker.input_q.put_nowait((frame.copy(), frame_num))
            except queue.Full:
                pass   # worker still busy, skip this frame

        # Get latest results from worker (non-blocking)
        try:
            last_results = worker.result_q.get_nowait()
            # Count new violations
            for d in last_results:
                if d["violations"] and d["plate"]:
                    total_viol += 1
        except queue.Empty:
            pass   # no new results yet, keep drawing old ones

        # ── Draw on display frame ────────────────────
        display = frame.copy()

        # Top bar
        cv2.rectangle(display, (0,0), (W,30), (15,35,55), -1)
        cv2.putText(display,
            f"Traffic Violation System  |  Frame:{frame_num}  "
            f"|  Violations:{total_viol}  |  Q=quit",
            (8, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.46, (255,255,255), 1)

        # Draw detection results
        draw_results(display, last_results, total_viol)

        # Bottom bar
        cv2.rectangle(display, (0, H-26), (W, H), (15,35,55), -1)
        cv2.putText(display,
            f"DB: {len(owners)} plates  |  "
            f"Match: {int(PLATE_MATCH_MIN*100)}%  |  "
            f"Detect every {DETECT_EVERY} frames  |  "
            f"Resize: {RESIZE_WIDTH}px",
            (8, H-8), cv2.FONT_HERSHEY_SIMPLEX, 0.38, (160,160,160), 1)

        cv2.imshow("Traffic Violation System  [Q=quit]", display)

        # waitKey controls playback speed - matches original video FPS
        key = cv2.waitKey(frame_delay) & 0xFF
        if key == ord('q'):
            break

    worker.stop()
    cap.release()
    cv2.destroyAllWindows()

    print(f"\n[DONE] Total violations: {total_viol}")
    print("  violation_images/ -> evidence photos")
    print("  challans/         -> PDF challans")
    print("  violations.xlsx   -> all records")


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else 0
    run(src)
